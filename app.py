"""
答题卡小题分匹配系统 - 主程序
将学生小题分叠加到扫描答题卡上，生成分析用图片
"""
from __future__ import annotations
import os
import re
import csv
import json
import sys
import uuid
import atexit
import signal
import shutil
import zipfile
import threading
import traceback
import multiprocessing
import concurrent.futures
import concurrent.futures.process  # 显式导入以确保 BrokenProcessPool 在打包版可用
from io import BytesIO
from pathlib import Path
from collections import defaultdict, Counter

from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from PIL import Image, ImageDraw, ImageFont, ImageEnhance
import numpy as np

# ---- 可选依赖 ----
try:
    from pyzbar import pyzbar
    HAS_PYZBAR = True
except Exception:
    # DLL 加载失败 / VC++ 缺失等场景会抛非 ImportError 的异常
    HAS_PYZBAR = False


def _check_pyzbar_functional() -> bool:
    """测试 pyzbar 是否能实际解码（仅 import 成功不代表可用，还需 VC++ 运行时）。
    解码失败时主动返回 False，避免静默丢失所有条码。"""
    if not HAS_PYZBAR:
        return False
    try:
        from PIL import Image as _PILImage
        test_img = _PILImage.new('L', (100, 100))
        pyzbar.decode(test_img)
        return True
    except Exception:
        return False

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except Exception:
    HAS_PYMUPDF = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:
    # 不仅捕获 ImportError：xlsx 链 (openpyxl -> xml -> pyexpat) 在打包后可能因 DLL 缺失抛非 ImportError
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except Exception:
    HAS_XLRD = False

try:
    from pypinyin import lazy_pinyin, Style
    HAS_PYPINYIN = True
except Exception:
    HAS_PYPINYIN = False

# ---- Flask 初始化 ----
app = Flask(__name__)
app.secret_key = os.urandom(24)
# 不限制上传大小（本地单机使用，磁盘空间就是上限）
app.config['MAX_CONTENT_LENGTH'] = None

# 兼容 PyInstaller 打包后的路径
IS_FROZEN = getattr(sys, 'frozen', False)
if IS_FROZEN:
    if hasattr(sys, '_MEIPASS'):
        BASE_DIR = Path(sys._MEIPASS)
    else:
        exe_dir = Path(sys.executable).parent
        internal_dir = exe_dir / '_internal'
        BASE_DIR = internal_dir if internal_dir.exists() else exe_dir
    # 打包版：所有 runtime 数据放到用户级 LOCALAPPDATA，不污染安装目录
    _local_appdata = os.environ.get('LOCALAPPDATA') or str(Path.home() / 'AppData' / 'Local')
    APP_USER_DIR = Path(_local_appdata) / '答题卡小题分匹配系统'
else:
    BASE_DIR = Path(__file__).parent
    # 开发版：仍放项目根目录，方便调试观察
    APP_USER_DIR = BASE_DIR

APP_USER_DIR.mkdir(parents=True, exist_ok=True)

DATA_DIR = APP_USER_DIR / 'data'
TEMPLATES_POS_DIR = APP_USER_DIR / 'templates_pos'
TEMPLATES_POS_DIR.mkdir(parents=True, exist_ok=True)

# 打包版首次运行：把内置 templates_pos 默认模板拷贝到用户目录
if IS_FROZEN:
    _bundled_tpos = BASE_DIR / 'templates_pos'
    if _bundled_tpos.exists() and not any(TEMPLATES_POS_DIR.glob('*.json')):
        for _p in _bundled_tpos.glob('*.json'):
            try:
                shutil.copy2(str(_p), str(TEMPLATES_POS_DIR / _p.name))
            except Exception:
                pass

app.template_folder = str(BASE_DIR / 'templates')


@app.errorhandler(413)
def request_entity_too_large(e):
    # MAX_CONTENT_LENGTH 已设为 None，正常情况下不会触发；保留兜底
    return jsonify({'error': '请求体过大，被底层服务器拒绝。请尝试分批上传。'}), 413


# 内存中存储 job 配置
jobs = {}

_cleaned = False
_shutdown_event = threading.Event()

# PDF 渲染持久化进程池（懒初始化，跨上传共用）
_pdf_executor: 'concurrent.futures.ProcessPoolExecutor | None' = None
_pdf_executor_workers = 0
_pdf_executor_lock = threading.Lock()
# 池失效计数：避免反复尝试已知失败的进程池；超过阈值后永久退化单进程
_pdf_executor_failures = 0
_PDF_EXECUTOR_MAX_FAILURES = 2


def _pdf_worker_warmup():
    """ProcessPoolExecutor 预热任务：让 worker 进程完成首次模块导入并执行函数，
    把潜在的"worker 启动失败"在用户提交真实工作前就暴露出来。
    返回 worker 进程 PID，用于日志。
    """
    import fitz  # 验证 worker 进程能导入 PyMuPDF（DLL 路径等问题会导致非 ImportError 异常）
    return os.getpid()


def _get_pdf_executor():
    """懒初始化持久化进程池。仅在 MainProcess 中创建，避免子进程内递归 spawn。

    首次创建池后会派发若干 warmup 任务并等待完成，确认所有 worker 能正常启动；
    若 warmup 失败（worker 启动错、DLL 缺失等），关闭池并返回 None，调用方走单进程。
    池失效（broken）时返回 None；连续失败超过阈值后永久退化单进程。
    """
    global _pdf_executor, _pdf_executor_workers, _pdf_executor_failures
    if multiprocessing.current_process().name != 'MainProcess':
        return None
    if _pdf_executor_failures >= _PDF_EXECUTOR_MAX_FAILURES:
        return None
    with _pdf_executor_lock:
        # 缓存池仍可用则直接返回；broken 则丢弃后重建
        if _pdf_executor is not None:
            if getattr(_pdf_executor, '_broken', False):
                try:
                    _pdf_executor.shutdown(wait=False)
                except Exception:
                    pass
                _pdf_executor = None
                _pdf_executor_workers = 0
            else:
                return _pdf_executor

        try:
            n = min(os.cpu_count() or 4, 8)
            pool = concurrent.futures.ProcessPoolExecutor(max_workers=n)
        except Exception as e:
            print(f'[pdf_pool] 启动失败，降级单进程: {e}')
            _pdf_executor = None
            _pdf_executor_workers = 0
            _pdf_executor_failures += 1
            return None

        # 预热：派发 n 个轻量任务，强制所有 worker 进程完成首次启动；
        # 任一 worker 启动失败会在 fut.result() 抛 BrokenProcessPool，
        # 比让用户上传 PDF 时才发现要早得多。
        try:
            warmup_futs = [pool.submit(_pdf_worker_warmup) for _ in range(n)]
            pids = []
            for fut in warmup_futs:
                pids.append(fut.result(timeout=60))
            print(f'[pdf_pool] PDF 渲染进程池已启动 + 预热完成，{n} 个 worker（PIDs={pids}）')
        except Exception as we:
            print(f'[pdf_pool] worker 预热失败（首次 PDF 将走单进程兜底）: {we}')
            try:
                try:
                    pool.shutdown(wait=False, cancel_futures=True)
                except TypeError:
                    pool.shutdown(wait=False)
            except Exception:
                pass
            _pdf_executor = None
            _pdf_executor_workers = 0
            _pdf_executor_failures += 1
            return None

        _pdf_executor = pool
        _pdf_executor_workers = n
    return _pdf_executor


def _reset_pdf_executor():
    """标记当前进程池失效（在 BrokenProcessPool 后调用），下次 _get_pdf_executor 自动重建。"""
    global _pdf_executor, _pdf_executor_workers
    with _pdf_executor_lock:
        if _pdf_executor is not None:
            try:
                try:
                    _pdf_executor.shutdown(wait=False, cancel_futures=True)
                except TypeError:
                    _pdf_executor.shutdown(wait=False)
            except Exception:
                pass
            _pdf_executor = None
            _pdf_executor_workers = 0


def _shutdown_pdf_executor():
    global _pdf_executor
    with _pdf_executor_lock:
        if _pdf_executor is not None:
            try:
                _pdf_executor.shutdown(wait=False, cancel_futures=True)
            except TypeError:
                _pdf_executor.shutdown(wait=False)
            except Exception:
                pass
            _pdf_executor = None


def cleanup_data():
    """清理 data 缓存目录"""
    data_dir = DATA_DIR
    try:
        if data_dir.exists():
            shutil.rmtree(str(data_dir), ignore_errors=True)
        data_dir.mkdir(parents=True, exist_ok=True)
        print(f'[init] 缓存目录已就绪: {data_dir}')
    except Exception as e:
        print(f'[init] 缓存目录初始化失败: {e}')


def _signal_handler(signum, frame):
    global _cleaned
    print(f'\n[signal] 收到信号 {signum}，正在退出...')
    _shutdown_event.set()
    if not _cleaned:
        _cleaned = True
        _shutdown_pdf_executor()
        cleanup_data()
    sys.exit(0)


def _atexit_cleanup():
    global _cleaned
    if not _cleaned:
        _cleaned = True
        _shutdown_event.set()
        _shutdown_pdf_executor()
        cleanup_data()


# 仅在主进程中执行清理和注册信号（Windows spawn 子进程重新导入本模块时跳过）
# 注：不能用 parent_process()——spawn 模式不会设置 _parent_process，它永远为 None。
if multiprocessing.current_process().name == 'MainProcess':
    cleanup_data()
    atexit.register(_atexit_cleanup)
    signal.signal(signal.SIGINT, _signal_handler)
    signal.signal(signal.SIGTERM, _signal_handler)


# ============================================================
#  工具函数
# ============================================================

def get_job_dir(job_id: str) -> Path:
    d = DATA_DIR / job_id
    d.mkdir(exist_ok=True)
    return d


# ============================================================
#  进度追踪（用于前端进度条）
# ============================================================

def _set_progress(job_id: str, task: str, current: int, total: int, message: str = '', **extra):
    """更新任务的进度信息"""
    job = jobs.get(job_id)
    if job:
        job['progress'] = {'task': task, 'current': current, 'total': total, 'message': message, **extra}


def _start_async_task(job_id: str, task_name: str, total: int, target_fn, *args):
    """在后台线程中运行任务，自动管理进度。target_fn(job_id, progress_cb, *args) 会在后台线程中被调用。"""
    job = jobs.get(job_id)
    if not job:
        return

    _set_progress(job_id, task_name, 0, total, '正在准备...')

    def progress_cb(current: int, message: str = ''):
        _set_progress(job_id, task_name, current, total, message)

    def worker():
        try:
            result = target_fn(job_id, progress_cb, *args)
            _set_progress(job_id, 'done', total, total, '完成', result=result)
        except Exception as e:
            traceback.print_exc()
            _set_progress(job_id, 'error', 0, total, str(e), error=str(e))

    t = threading.Thread(target=worker, daemon=True)
    t.start()


@app.route('/api/progress/<job_id>')
def get_progress(job_id):
    """获取任务进度"""
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400
    progress = jobs[job_id].get('progress', {'task': None, 'current': 0, 'total': 0, 'message': ''})
    return jsonify(progress)


@app.route('/api/image-list/<job_id>')
def get_image_list(job_id):
    """获取当前已上传的图片列表"""
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400
    job_dir = get_job_dir(job_id)
    img_dir = job_dir / 'images'
    if not img_dir.exists():
        return jsonify({'images': [], 'total': 0})
    all_images = sorted([p.name for p in img_dir.iterdir() if _is_image_file(p)])
    jobs[job_id]['image_count'] = len(all_images)
    return jsonify({'images': all_images, 'total': len(all_images)})


def find_chinese_font(font_size: int) -> ImageFont.FreeTypeFont | None:
    """查找支持中文的字体"""
    font_paths = [
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/msyh.ttf",
        "C:/Windows/Fonts/msyhbd.ttf",
        "C:/Windows/Fonts/simsun.ttc",
        "C:/Windows/Fonts/simkai.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttf",
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ]
    for fp in font_paths:
        try:
            return ImageFont.truetype(fp, font_size)
        except (IOError, OSError):
            continue
    try:
        return ImageFont.load_default()
    except Exception:
        return None


def get_xlsx_sheet_names(filepath: str) -> list[str]:
    """获取 xlsx 文件中所有工作表名称"""
    if not HAS_OPENPYXL:
        return []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        pass
    # read_only 模式可能因打包后缺少子模块而失败，回退到普通模式
    try:
        wb = openpyxl.load_workbook(filepath)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        print(f'[WARN] get_xlsx_sheet_names failed: {e}')
        return []


def parse_score_file(filepath: str, sheet_name: str | None = None) -> list[list[str]]:
    """解析成绩文件 (CSV / XLSX / XLS)。sheet_name 仅对 xlsx 有效，默认使用活动工作表。"""
    ext = Path(filepath).suffix.lower()
    rows = []

    if ext == '.csv':
        for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb18030', 'latin-1']:
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    reader = csv.reader(f)
                    rows = [list(row) for row in reader]
                break
            except (UnicodeDecodeError, UnicodeError):
                continue

    elif ext in ('.xlsx', '.xls'):
        if ext == '.xlsx' and HAS_OPENPYXL:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            rows = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    if cell.value is None:
                        row_data.append('')
                    else:
                        row_data.append(str(cell.value))
                rows.append(row_data)
            wb.close()
        elif ext == '.xls' and HAS_XLRD:
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_index(0)
            rows = []
            for r in range(ws.nrows):
                rows.append([str(ws.cell_value(r, c)) if ws.cell_value(r, c) != '' else '' for c in range(ws.ncols)])
        else:
            for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']:
                try:
                    with open(filepath, 'r', encoding=enc) as f:
                        reader = csv.reader(f)
                        rows = [list(row) for row in reader]
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue

    rows = [[c.strip() for c in row] for row in rows]
    return rows


def read_barcodes_from_image(image: Image.Image, region: dict | None = None) -> list[dict]:
    """读取图片中的条码。region 为 {'x1','y1','x2','y2'} 百分比裁切区域（可选）。"""
    if not HAS_PYZBAR:
        return []
    try:
        if region:
            w, h = image.size
            x1 = int(w * region['x1'] / 100)
            y1 = int(h * region['y1'] / 100)
            x2 = int(w * region['x2'] / 100)
            y2 = int(h * region['y2'] / 100)
            x1, x2 = sorted([max(0, x1), min(w, x2)])
            y1, y2 = sorted([max(0, y1), min(h, y2)])
            if x2 - x1 > 10 and y2 - y1 > 10:
                image = image.crop((x1, y1, x2, y2))
        gray = image.convert('L')
        barcodes = pyzbar.decode(gray)
        results = []
        for bc in barcodes:
            results.append({
                'type': bc.type,
                'data': bc.data.decode('utf-8', errors='ignore'),
                'rect': {
                    'left': bc.rect.left,
                    'top': bc.rect.top,
                    'width': bc.rect.width,
                    'height': bc.rect.height,
                }
            })
        return results
    except Exception:
        return []


def extract_digits_from_filename(filename: str) -> str:
    """从文件名提取连续数字作为可能的考号"""
    digits = re.findall(r'\d+', filename)
    if digits:
        return max(digits, key=len)
    return ''


def convert_pdf_to_images(pdf_path: str, output_dir: str, prefix: str = '') -> list[str]:
    """将 PDF 的每一页转为 JPG 图片。prefix 用于区分不同 PDF 的页面文件。"""
    if not HAS_PYMUPDF:
        raise RuntimeError("需要安装 PyMuPDF 来处理 PDF 文件: pip install pymupdf")
    if not prefix:
        prefix = Path(pdf_path).stem
    # 清理前缀中的特殊字符
    prefix = re.sub(r'[^\w\-]', '_', prefix)
    doc = fitz.open(pdf_path)
    image_paths = []
    total = len(doc)
    for page_num in range(total):
        page = doc[page_num]
        mat = fitz.Matrix(300 / 72, 300 / 72)
        pix = page.get_pixmap(matrix=mat)
        img_path = os.path.join(output_dir, f'{prefix}_{page_num + 1:04d}.jpg')
        pix.save(img_path)
        image_paths.append(img_path)
    doc.close()
    return image_paths


def add_scores_to_image(
    image_path: str,
    x_percentages: list[float],
    y_percentages: list[float],
    texts: list[str],
    font_size: int,
    save_path: str,
    font: ImageFont.FreeTypeFont | None = None,
):
    """在图片指定百分比位置添加小题分文字"""
    if len({len(x_percentages), len(y_percentages), len(texts)}) != 1:
        raise ValueError("坐标列表与文字列表长度必须一致")

    image = Image.open(image_path).convert('RGB')
    width, height = image.size
    draw = ImageDraw.Draw(image)

    if font is None:
        font = find_chinese_font(font_size)
    if font is None:
        raise RuntimeError("无法加载字体")

    for x_pct, y_pct, text in zip(x_percentages, y_percentages, texts):
        x = int(width * x_pct / 100)
        y = int(height * y_pct / 100)

        try:
            parts = text.split('/')
            stu_score = parts[0].strip()
            full_score = parts[1].strip()
            if stu_score == full_score:
                color = (104, 151, 187)  # 蓝色-满分
            else:
                color = (255, 0, 0)      # 红色-扣分
        except Exception:
            color = (255, 0, 0)

        draw.text((x, y), text, font=font, fill=color, anchor='mm')

    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    image.save(save_path, 'JPEG', quality=95)


def image_info(image_path: str) -> dict:
    """获取图片基本信息"""
    try:
        img = Image.open(image_path)
        w, h = img.size
        return {'width': w, 'height': h, 'format': img.format, 'mode': img.mode}
    except Exception:
        return {'width': 0, 'height': 0, 'format': '', 'mode': ''}


def _is_image_file(p: Path) -> bool:
    return p.suffix.lower() in ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif')


def _build_student_output_names(students: list[dict]) -> dict:
    """
    为学生构建去重后的输出文件名（不含扩展名）。
    返回 {student_id: output_base_name}

    规则：
    - 优先使用学生姓名
    - 同班重名：后加 a, b, c... 后缀
    """
    # 按班级分组
    class_groups = defaultdict(list)
    for s in students:
        cls = s.get('class_name', 'unknown')
        class_groups[cls].append(s)

    name_map = {}
    for cls, stu_list in class_groups.items():
        name_counts = defaultdict(list)
        for s in stu_list:
            name = s.get('student_name', s['student_id'])
            name_counts[name].append(s['student_id'])

        # 处理重名
        name_suffix = {}
        for name, ids in name_counts.items():
            if len(ids) == 1:
                name_suffix[ids[0]] = name
            else:
                for i, sid in enumerate(ids):
                    suffix = chr(ord('a') + i)  # a, b, c, ...
                    name_suffix[sid] = f'{name}{suffix}'

        for s in stu_list:
            name_map[s['student_id']] = name_suffix[s['student_id']]

    return name_map


def _get_pinyin(text: str) -> tuple[str, str]:
    """返回中文文本的拼音首字母和全拼，如 ('zs', 'zhangsan')"""
    if not HAS_PYPINYIN or not text:
        return '', ''
    try:
        initials = lazy_pinyin(text, style=Style.FIRST_LETTER)
        full = lazy_pinyin(text, style=Style.NORMAL)
        return ''.join(initials).lower(), ''.join(full).lower()
    except Exception:
        return '', ''


def _enrich_student_list(students: list[dict]) -> list[dict]:
    """为学生列表添加拼音搜索字段"""
    result = []
    for s in students:
        name = s.get('student_name', '')
        initials, full_py = _get_pinyin(name)
        item = {
            'student_id': s['student_id'],
            'student_name': name,
            'class_name': s.get('class_name', ''),
            'py_initials': initials,
            'py_full': full_py,
        }
        result.append(item)
    return result


# ============================================================
#  API 路由
# ============================================================

@app.route('/')
def index():
    """主页面"""
    return render_template('index.html')


@app.route('/api/init-job', methods=['POST'])
def init_job():
    """初始化新的处理任务"""
    job_id = uuid.uuid4().hex[:12]
    job_dir = get_job_dir(job_id)
    (job_dir / 'scores').mkdir(exist_ok=True)
    (job_dir / 'images').mkdir(exist_ok=True)
    (job_dir / 'output').mkdir(exist_ok=True)
    (job_dir / 'unmatched').mkdir(exist_ok=True)

    jobs[job_id] = {
        'id': job_id,
        'score_file': None,
        'score_rows': [],
        'score_config': None,
        'image_count': 0,
        'positions_front': [],
        'positions_back': [],
        'manual_matches': {},       # {image_idx_or_pair_idx: student_id}
        'barcode_scan_done': False,
        'processed': 0,
        'matched': 0,
        'unmatched_students': [],
        'unmatched_images': [],
        'output_files': [],
        'status': 'init',
    }
    return jsonify({'job_id': job_id})


@app.route('/api/upload-scores', methods=['POST'])
def upload_scores():
    """上传成绩文件并返回预览。支持 xlsx 工作表选择。"""
    job_id = request.form.get('job_id', '')
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    file = request.files.get('file')
    sheet_name = request.form.get('sheet_name', '') or None

    # 如果传了 sheet_name，说明是切换工作表，使用已有文件
    if sheet_name:
        job = jobs[job_id]
        score_file = job.get('score_file', '')
        if not score_file:
            return jsonify({'error': '请先上传成绩文件'}), 400
        try:
            rows = parse_score_file(score_file, sheet_name=sheet_name)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {str(e)}'}), 400
        if not rows:
            return jsonify({'error': f'工作表 "{sheet_name}" 中未找到任何数据'}), 400
        job['score_rows'] = rows
        job['score_config'] = None
        job['score_sheet'] = sheet_name
        preview = rows[:20]
        max_cols = max(len(r) for r in preview) if preview else 0
        return jsonify({
            'total_rows': len(rows),
            'preview_rows': len(preview),
            'max_cols': max_cols,
            'preview': preview,
            'headers': [f'第{i+1}列' for i in range(max_cols)],
            'active_sheet': sheet_name,
        })

    if not file:
        return jsonify({'error': '未选择文件'}), 400

    ext = Path(file.filename).suffix.lower()
    if ext not in ('.csv', '.xls', '.xlsx'):
        return jsonify({'error': f'不支持的文件格式: {ext}，仅支持 CSV / XLS / XLSX'}), 400

    job_dir = get_job_dir(job_id)
    (job_dir / 'scores').mkdir(exist_ok=True)
    save_path = job_dir / 'scores' / f'score{ext}'
    file.save(str(save_path))

    try:
        rows = parse_score_file(str(save_path))
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400

    # 获取 xlsx 的工作表列表（用于前端切换）
    sheet_names = []
    if ext == '.xlsx':
        sheet_names = get_xlsx_sheet_names(str(save_path))

    if not rows and not sheet_names:
        return jsonify({'error': '文件中未找到任何数据'}), 400

    if not rows and sheet_names:
        # 活动工作表为空，但有其他工作表可选，尝试第一个
        for sn in sheet_names:
            rows = parse_score_file(str(save_path), sheet_name=sn)
            if rows:
                sheet_name = sn
                break
        if not rows:
            return jsonify({'error': '所有工作表中均未找到数据'}), 400
    else:
        # 获取实际使用的活动工作表名
        sheet_name = None
        if ext == '.xlsx' and HAS_OPENPYXL:
            try:
                wb = openpyxl.load_workbook(str(save_path), data_only=True)
                sheet_name = wb.active.title
                wb.close()
            except Exception:
                pass

    jobs[job_id]['score_file'] = str(save_path)
    jobs[job_id]['score_rows'] = rows
    jobs[job_id]['score_config'] = None
    jobs[job_id]['score_sheet'] = sheet_name

    preview = rows[:20]
    max_cols = max(len(r) for r in preview) if preview else 0

    return jsonify({
        'total_rows': len(rows),
        'preview_rows': len(preview),
        'max_cols': max_cols,
        'preview': preview,
        'headers': [f'第{i+1}列' for i in range(max_cols)],
        'sheet_names': sheet_names,
        'active_sheet': sheet_name,
    })


# ---- 后台任务 worker 函数 ----

def _render_single_page_worker(pdf_path: str, page_idx: int, out_path: str) -> dict:
    """ProcessPoolExecutor worker：渲染 PDF 的单页。必须是模块顶层函数才能被 pickle。
    每次调用独立 open/close PDF（PDF 解析很快，换取每页一个 future 的细粒度进度）。
    不能访问父进程的 _shutdown_event / jobs 等模块状态。
    """
    try:
        import fitz as _fitz
    except Exception:
        # DLL 加载失败等场景会抛非 ImportError 异常（尤其在 PyInstaller worker 进程中）
        return {'ok': False, 'error': 'fitz/PyMuPDF 在 worker 进程中不可用'}

    doc = None
    try:
        doc = _fitz.open(pdf_path)
        page = doc[page_idx]
        mat = _fitz.Matrix(300 / 72, 300 / 72)
        pix = page.get_pixmap(matrix=mat)
        pix.save(out_path)
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': f'页 {page_idx + 1}: {e}'}
    finally:
        if doc is not None:
            try:
                doc.close()
            except Exception:
                pass


def _do_pdf_convert(job_id: str, progress_cb, pdf_info_list: list[dict], img_dir: str):
    """后台转换 PDF 文件为图片。

    用持久化 ProcessPoolExecutor 把单个 PDF 的页面分发到多个进程并行渲染，
    绕过 GIL 实现真正的 CPU 并行。多个 PDF 顺序处理（前端按 PDF 逐个提交）。

    输出命名：`{全局序号:05d}_{原 PDF stem}.jpg`
    全局序号 = 转换开始前 images/ 已有图片数 + 累计已转换页 + 当前页序 + 1
    多次上传 / 同名 PDF 重复上传都不会覆盖；自然按上传顺序排序。
    """
    total_pages = sum(p['pages'] for p in pdf_info_list)
    if total_pages <= 0:
        return {'converted_pages': 0, 'errors': [], 'pdf_count': 0}

    img_dir_path = Path(img_dir)
    img_dir_path.mkdir(parents=True, exist_ok=True)
    base_seq = len([p for p in img_dir_path.iterdir() if _is_image_file(p)])
    converted_total = 0
    errors: list[str] = []

    # 真实进度：派发到 worker 前先给一条更准确的消息
    progress_cb(0, f'读取 PDF，准备 {total_pages} 页渲染任务...')

    MP_MIN_PAGES = 2  # 单页 PDF 不值得起进程间通信

    for info in pdf_info_list:
        if _shutdown_event.is_set():
            break
        pages = info['pages']
        if pages <= 0:
            continue

        # 给本 PDF 每一页分配输出路径（全局序号在所有上传间唯一）
        page_assignments = []
        for p_idx in range(pages):
            seq = base_seq + converted_total + p_idx + 1
            out_name = f'{seq:05d}_{info["prefix"]}.jpg'
            page_assignments.append((p_idx, os.path.join(img_dir, out_name)))

        executor = _get_pdf_executor() if pages >= MP_MIN_PAGES else None

        if executor is None:
            # 单进程降级路径（单页 PDF 或进程池不可用 / 预热失败）
            progress_cb(converted_total, f'{info["prefix"]}：单进程渲染中... ({converted_total}/{total_pages})')
            for p_idx, out_path in page_assignments:
                if _shutdown_event.is_set():
                    break
                r = _render_single_page_worker(info['path'], p_idx, out_path)
                if r.get('ok'):
                    converted_total += 1
                else:
                    errors.append(f'{info["prefix"]}: {r.get("error", "未知错误")}')
                progress_cb(converted_total, f'{info["prefix"]}：渲染中 {converted_total}/{total_pages}')
        else:
            # 多进程：每页一个 future，进度按页粒度更新
            progress_cb(converted_total, f'{info["prefix"]}：派发 {pages} 页到 {_pdf_executor_workers} 个进程...')
            futures = []
            pool_broken = False
            try:
                for p_idx, out_path in page_assignments:
                    futures.append(executor.submit(_render_single_page_worker, info['path'], p_idx, out_path))
                # 派发完毕，告诉用户 worker 即将/正在启动
                progress_cb(converted_total, f'{info["prefix"]}：worker 进程启动中... ({converted_total}/{total_pages})')

                for fut in concurrent.futures.as_completed(futures):
                    if pool_broken:
                        # 池已损坏，后续 future 都会抛同样错；跳过，下面单进程兜底
                        try:
                            fut.cancel()
                        except Exception:
                            pass
                        continue
                    try:
                        r = fut.result()
                        if r.get('ok'):
                            converted_total += 1
                        else:
                            errors.append(f'{info["prefix"]}: {r.get("error", "未知错误")}')
                        progress_cb(converted_total, f'{info["prefix"]}：渲染中 {converted_total}/{total_pages}')
                    except concurrent.futures.process.BrokenProcessPool as bpe:
                        pool_broken = True
                        errors.append(f'{info["prefix"]}: 进程池崩溃，本 PDF 剩余页转单进程渲染: {bpe}')
                        progress_cb(converted_total, f'{info["prefix"]}：进程池崩溃，单进程兜底中...')
                        _reset_pdf_executor()
                    except Exception as e:
                        errors.append(f'{info["prefix"]} worker 异常: {e}')
                        traceback.print_exc()
            except concurrent.futures.process.BrokenProcessPool as bpe:
                # 提交阶段就抛 BrokenProcessPool（如池在 submit 中被发现 broken）
                pool_broken = True
                errors.append(f'{info["prefix"]}: 进程池在提交时崩溃，降级单进程: {bpe}')
                _reset_pdf_executor()
            except Exception as e:
                # 提交失败（如进程池已关闭等）：降级到当前进程
                pool_broken = True
                errors.append(f'{info["prefix"]}: 进程池提交失败，降级单进程: {e}')
                _reset_pdf_executor()

            if pool_broken:
                # 单进程兜底：把还没渲染成功的页（输出文件不存在）重渲一遍
                for p_idx, out_path in page_assignments:
                    if _shutdown_event.is_set():
                        break
                    if Path(out_path).exists():
                        continue
                    r = _render_single_page_worker(info['path'], p_idx, out_path)
                    if r.get('ok'):
                        converted_total += 1
                    else:
                        errors.append(f'{info["prefix"]}: 单进程兜底失败: {r.get("error", "未知错误")}')
                    progress_cb(converted_total, f'{info["prefix"]}：单进程兜底 {converted_total}/{total_pages}')

    # 清理原始 PDF 文件（已转为 JPG）
    for info in pdf_info_list:
        try:
            os.remove(info['path'])
        except Exception:
            pass

    job = jobs.get(job_id)
    if job:
        job['image_count'] = len([p for p in img_dir_path.iterdir() if _is_image_file(p)])

    return {
        'converted_pages': converted_total,
        'pdf_count': len(pdf_info_list),
        'errors': errors,
    }


def _do_barcode_scan(job_id: str, progress_cb, double_sided: bool, use_barcode: bool, barcode_region: dict | None):
    """后台扫描条码"""
    job = jobs.get(job_id)
    if not job:
        return
    config = job.get('score_config')
    rows = job['score_rows']
    parsed = _parse_scores_with_config(rows, config)
    student_map = {s['student_id']: s for s in parsed['students']}
    student_list = _enrich_student_list(parsed['students'])

    job_dir = get_job_dir(job_id)
    img_dir = job_dir / 'images'
    image_files = sorted([p for p in img_dir.iterdir() if _is_image_file(p)])

    matched = []
    unmatched = []
    all_barcode_reads = []

    if double_sided:
        pair_count = len(image_files) // 2
        progress_cb(0, f'开始扫描 {pair_count} 对试卷...')
        for pair_idx in range(pair_count):
            front_img = image_files[pair_idx * 2]
            back_img = image_files[pair_idx * 2 + 1]
            student_id = None
            barcode_raw = ''
            manual_key = str(pair_idx)
            if manual_key in job.get('manual_matches', {}):
                student_id = job['manual_matches'][manual_key]
            elif use_barcode and HAS_PYZBAR:
                try:
                    img_obj = Image.open(front_img)
                    barcodes = read_barcodes_from_image(img_obj, barcode_region)
                    if barcodes:
                        barcode_raw = barcodes[0]['data'].strip()
                        student_id = barcode_raw
                        all_barcode_reads.append({
                            'barcode_data': barcode_raw,
                            'image': front_img.name,
                            'pair_idx': pair_idx,
                            'matched_student': student_id if student_id in student_map else None,
                        })
                except Exception:
                    pass
            if not student_id:
                student_id = extract_digits_from_filename(front_img.stem)
            if student_id and student_id in student_map:
                s = student_map[student_id]
                matched.append({
                    'student_id': student_id,
                    'student_name': s.get('student_name', ''),
                    'class_name': s.get('class_name', ''),
                    'front_image': front_img.name,
                    'back_image': back_img.name,
                    'pair_idx': pair_idx,
                })
            else:
                unmatched.append({
                    'front_image': front_img.name,
                    'back_image': back_img.name,
                    'pair_idx': pair_idx,
                    'guessed_id': student_id or '',
                })
            progress_cb(pair_idx + 1, f'扫描条码: {pair_idx + 1}/{pair_count}')
        if len(image_files) % 2 == 1:
            last_img = image_files[-1]
            unmatched.append({
                'front_image': last_img.name,
                'back_image': None,
                'pair_idx': pair_count,
                'guessed_id': extract_digits_from_filename(last_img.stem),
            })
    else:
        total = len(image_files)
        progress_cb(0, f'开始扫描 {total} 张试卷...')
        for idx, img_file in enumerate(image_files):
            student_id = None
            barcode_raw = ''
            manual_key = str(idx)
            if manual_key in job.get('manual_matches', {}):
                student_id = job['manual_matches'][manual_key]
            elif use_barcode and HAS_PYZBAR:
                try:
                    img_obj = Image.open(img_file)
                    barcodes = read_barcodes_from_image(img_obj, barcode_region)
                    if barcodes:
                        barcode_raw = barcodes[0]['data'].strip()
                        student_id = barcode_raw
                        all_barcode_reads.append({
                            'barcode_data': barcode_raw,
                            'image': img_file.name,
                            'idx': idx,
                            'matched_student': student_id if student_id in student_map else None,
                        })
                except Exception:
                    pass
            if not student_id:
                student_id = extract_digits_from_filename(img_file.stem)
            if student_id and student_id in student_map:
                s = student_map[student_id]
                matched.append({
                    'student_id': student_id,
                    'student_name': s.get('student_name', ''),
                    'class_name': s.get('class_name', ''),
                    'image': img_file.name,
                    'idx': idx,
                })
            else:
                unmatched.append({
                    'image': img_file.name,
                    'idx': idx,
                    'guessed_id': student_id or '',
                })
            progress_cb(idx + 1, f'扫描条码: {idx + 1}/{total}')

    # 重复条码检测
    dup_barcodes = []
    if all_barcode_reads:
        barcode_counts = Counter(r['barcode_data'] for r in all_barcode_reads if r['barcode_data'])
        dup_values = {v for v, c in barcode_counts.items() if c > 1}
        if dup_values:
            dup_barcodes = [r for r in all_barcode_reads if r['barcode_data'] in dup_values]

    matched_ids = {m['student_id'] for m in matched}
    unmatched_students = []
    for sid, s in student_map.items():
        if sid not in matched_ids:
            unmatched_students.append({
                'student_id': sid,
                'student_name': s.get('student_name', ''),
                'class_name': s.get('class_name', ''),
            })

    job['barcode_scan_done'] = True

    return {
        'double_sided': double_sided,
        'matched': matched,
        'unmatched': unmatched,
        'unmatched_students': unmatched_students,
        'student_list': student_list,
        'total_images': len(image_files),
        'dup_barcodes': dup_barcodes,
    }


def _do_process(job_id: str, progress_cb, double_sided: bool, use_barcode: bool,
                font_size: int, create_class_folders: bool):
    """后台处理：匹配和分数叠加"""
    job = jobs.get(job_id)
    if not job:
        return
    config = job.get('score_config')
    rows = job['score_rows']
    parsed = _parse_scores_with_config(rows, config)
    full_scores = parsed['full_scores']
    students = parsed['students']
    student_map = {s['student_id']: s for s in students}
    output_names = _build_student_output_names(students)

    job_dir = get_job_dir(job_id)
    img_dir = job_dir / 'images'
    output_dir = job_dir / 'output'
    unmatched_dir = job_dir / 'unmatched'
    output_dir.mkdir(exist_ok=True)
    unmatched_dir.mkdir(exist_ok=True)

    image_files = sorted([p for p in img_dir.iterdir() if _is_image_file(p)])
    manual_matches = job.get('manual_matches', {})
    barcode_region = job.get('barcode_region')
    positions_front = job.get('positions_front', [])
    positions_back = job.get('positions_back', [])

    font = find_chinese_font(font_size)
    if font is None:
        raise RuntimeError('无法加载字体文件')

    class_dirs: set[str] = set()
    output_files = []
    matched_ids = set()
    matched_count = 0
    unmatched_images: list[str] = []

    def _make_score_text(score_val, idx):
        fs = full_scores[idx] if idx < len(full_scores) else '?'
        return f'{score_val}/{fs}'

    if double_sided:
        pair_count = len(image_files) // 2
        progress_cb(0, f'开始处理 {pair_count} 对试卷...')
        for pair_idx in range(pair_count):
            front_img = image_files[pair_idx * 2]
            back_img = image_files[pair_idx * 2 + 1]
            front_path = str(front_img)
            back_path = str(back_img)
            student_id = None

            manual_key = str(pair_idx)
            if manual_key in manual_matches:
                student_id = manual_matches[manual_key]
            elif use_barcode and HAS_PYZBAR:
                try:
                    img_obj = Image.open(front_path)
                    barcodes = read_barcodes_from_image(img_obj, barcode_region)
                    if barcodes:
                        student_id = barcodes[0]['data'].strip()
                except Exception:
                    pass
            if not student_id:
                student_id = extract_digits_from_filename(front_img.stem)

            if not student_id or student_id not in student_map:
                shutil.copy(front_path, unmatched_dir / front_img.name)
                shutil.copy(back_path, unmatched_dir / back_img.name)
                unmatched_images.append(front_img.name)
                progress_cb(pair_idx + 1, f'处理: {pair_idx + 1}/{pair_count}')
                continue

            student = student_map[student_id]
            student_scores = student['scores']
            class_name = student.get('class_name', 'unknown')
            out_name = output_names.get(student_id, student_id)

            if create_class_folders:
                save_dir = output_dir / class_name
            else:
                save_dir = output_dir
            save_dir.mkdir(exist_ok=True)
            class_dirs.add(class_name)

            n_front = len(positions_front)
            front_texts = [_make_score_text(student_scores[j], j) for j in range(min(n_front, len(student_scores)))]
            while len(front_texts) < n_front:
                front_texts.append('?/?')
            out_front = save_dir / f'{out_name}1.jpg'
            add_scores_to_image(front_path, [p['x'] for p in positions_front],
                                [p['y'] for p in positions_front], front_texts, font_size, str(out_front), font)
            output_files.append(str(out_front))

            n_back = len(positions_back)
            back_scores = student_scores[n_front:n_front + n_back]
            back_texts = [_make_score_text(back_scores[j], n_front + j) for j in range(min(n_back, len(back_scores)))]
            while len(back_texts) < n_back:
                back_texts.append('?/?')
            out_back = save_dir / f'{out_name}2.jpg'
            add_scores_to_image(back_path, [p['x'] for p in positions_back],
                                [p['y'] for p in positions_back], back_texts, font_size, str(out_back), font)
            output_files.append(str(out_back))
            matched_ids.add(student_id)
            matched_count += 1
            progress_cb(pair_idx + 1, f'处理: {pair_idx + 1}/{pair_count}')

        if len(image_files) % 2 == 1:
            last_img = image_files[-1]
            shutil.copy(str(last_img), unmatched_dir / last_img.name)
            unmatched_images.append(last_img.name)
    else:
        total = len(image_files)
        progress_cb(0, f'开始处理 {total} 张试卷...')
        for idx, img_file in enumerate(image_files):
            img_path = str(img_file)
            student_id = None
            manual_key = str(idx)
            if manual_key in manual_matches:
                student_id = manual_matches[manual_key]
            elif use_barcode and HAS_PYZBAR:
                try:
                    img_obj = Image.open(img_path)
                    barcodes = read_barcodes_from_image(img_obj, barcode_region)
                    if barcodes:
                        student_id = barcodes[0]['data'].strip()
                except Exception:
                    pass
            if not student_id:
                student_id = extract_digits_from_filename(img_file.stem)

            if not student_id or student_id not in student_map:
                shutil.copy(img_path, unmatched_dir / img_file.name)
                unmatched_images.append(img_file.name)
                progress_cb(idx + 1, f'处理: {idx + 1}/{total}')
                continue

            student = student_map[student_id]
            student_scores = student['scores']
            class_name = student.get('class_name', 'unknown')
            out_name = output_names.get(student_id, student_id)

            if create_class_folders:
                save_dir = output_dir / class_name
            else:
                save_dir = output_dir
            save_dir.mkdir(exist_ok=True)
            class_dirs.add(class_name)

            n_pos = len(positions_front)
            texts = [_make_score_text(student_scores[j], j) for j in range(min(n_pos, len(student_scores)))]
            while len(texts) < n_pos:
                texts.append('?/?')
            out_path_obj = save_dir / f'{out_name}.jpg'
            add_scores_to_image(img_path, [p['x'] for p in positions_front],
                                [p['y'] for p in positions_front], texts, font_size, str(out_path_obj), font)
            output_files.append(str(out_path_obj))
            matched_ids.add(student_id)
            matched_count += 1
            progress_cb(idx + 1, f'处理: {idx + 1}/{total}')

    unmatched_students = []
    for sid, student in student_map.items():
        if sid not in matched_ids:
            unmatched_students.append({
                'student_id': sid,
                'student_name': student.get('student_name', ''),
                'class_name': student.get('class_name', ''),
            })

    return {
        'output_files': output_files,
        'matched_count': matched_count,
        'matched_ids': list(matched_ids),
        'unmatched_images': unmatched_images,
        'unmatched_students': unmatched_students,
        'total_images': len(image_files),
        'class_dirs': sorted(class_dirs),
    }


@app.route('/api/upload-images', methods=['POST'])
def upload_images():
    """上传答题卡图片"""
    job_id = request.form.get('job_id', '')
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '未选择文件'}), 400

    job_dir = get_job_dir(job_id)
    img_dir = job_dir / 'images'
    img_dir.mkdir(exist_ok=True)
    supported_exts = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.pdf'}

    saved = []
    errors = []
    pdf_list = []  # [(path, prefix, page_count)]

    for f in files:
        if not f.filename:
            continue
        ext = Path(f.filename).suffix.lower()
        if ext not in supported_exts:
            errors.append(f'{f.filename}: 不支持的格式')
            continue

        safe_name = secure_filename(f.filename)
        save_path = img_dir / safe_name
        counter = 1
        while save_path.exists():
            stem = Path(safe_name).stem
            save_path = img_dir / f'{stem}_{counter}{ext}'
            counter += 1
        f.save(str(save_path))

        if ext == '.pdf':
            try:
                pdf_prefix = Path(save_path).stem
                doc = fitz.open(str(save_path))
                page_count = len(doc)
                doc.close()
                pdf_list.append({'path': str(save_path), 'prefix': pdf_prefix, 'pages': page_count})
                saved.append(save_path.name)
            except Exception as e:
                errors.append(f'{f.filename}: PDF读取失败 - {str(e)}')
        else:
            saved.append(save_path.name)

    # 统计当前图片（不含PDF页面）
    all_images = sorted([p.name for p in img_dir.iterdir() if _is_image_file(p)])
    jobs[job_id]['image_count'] = len(all_images)

    total_pdf_pages = sum(p['pages'] for p in pdf_list)

    if pdf_list:
        _start_async_task(job_id, 'pdf_convert', total_pdf_pages, _do_pdf_convert, pdf_list, str(img_dir))

    return jsonify({
        'saved': saved,
        'total_images': len(all_images),
        'images': all_images,
        'errors': errors,
        'pdf_pending': total_pdf_pages > 0,
        'pdf_total_pages': total_pdf_pages,
    })


@app.route('/api/config-scores', methods=['POST'])
def config_scores():
    """配置成绩文件的解析参数"""
    job_id = request.json.get('job_id', '')
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    config = {
        'skip_rows': request.json.get('skip_rows', 0),
        'id_column': request.json.get('id_column', 0),
        'name_column': request.json.get('name_column', -1),
        'full_score_row': request.json.get('full_score_row', 0),
        'score_start_column': request.json.get('score_start_column', 5),
        'score_end_column': request.json.get('score_end_column', -1),
        'class_column': request.json.get('class_column', -1),
    }

    rows = jobs[job_id]['score_rows']
    if not rows:
        return jsonify({'error': '请先上传成绩文件'}), 400

    try:
        parsed = _parse_scores_with_config(rows, config)
    except Exception as e:
        return jsonify({'error': f'配置验证失败: {str(e)}'}), 400

    jobs[job_id]['score_config'] = config

    return jsonify({
        'student_count': len(parsed['students']),
        'question_count': len(parsed['full_scores']),
        'full_scores': parsed['full_scores'],
        'sample_students': parsed['students'][:5],
    })


def _parse_scores_with_config(rows: list[list[str]], config: dict) -> dict:
    """根据配置解析成绩数据"""
    skip = config['skip_rows']
    id_col = config['id_column']
    name_col = config.get('name_column', -1)
    full_row = config['full_score_row']
    score_start = config['score_start_column']
    score_end = config['score_end_column']
    class_col = config['class_column']

    if score_end == -1 or score_end is None:
        score_end = max(len(r) for r in rows)

    # 获取满分行
    if full_row < len(rows):
        full_scores = rows[full_row][score_start:score_end]
    else:
        full_scores = []

    # 解析学生数据
    students = []
    for i, row in enumerate(rows):
        if i < skip or i == full_row:
            continue
        if id_col >= len(row) or not row[id_col].strip():
            continue

        student = {
            'row_index': i,
            'student_id': row[id_col].strip(),
            'scores': row[score_start:score_end] if score_start < len(row) else [],
        }

        # 提取姓名
        if name_col >= 0 and name_col < len(row):
            student['student_name'] = row[name_col].strip()
        else:
            student['student_name'] = student['student_id']

        if class_col >= 0 and class_col < len(row):
            student['class_name'] = row[class_col].strip()

        students.append(student)

    return {
        'full_scores': full_scores,
        'students': students,
    }


# ============================================================
#  条码扫描 & 人工匹配
# ============================================================

@app.route('/api/scan-barcodes', methods=['POST'])
def scan_barcodes():
    """扫描所有已上传图片的条码。支持 async_mode 后台执行以显示进度条。"""
    job_id = request.json.get('job_id', '')
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    job = jobs[job_id]
    config = job.get('score_config')
    if not config:
        return jsonify({'error': '请先配置成绩文件'}), 400

    double_sided = request.json.get('double_sided', False)
    use_barcode = request.json.get('use_barcode', True)
    barcode_region = request.json.get('barcode_region', None)
    async_mode = request.json.get('async', True)

    if barcode_region:
        job['barcode_region'] = barcode_region

    # 计算总工作量
    job_dir = get_job_dir(job_id)
    img_dir = job_dir / 'images'
    image_files = sorted([p for p in img_dir.iterdir() if _is_image_file(p)])
    total = len(image_files) // 2 if double_sided else len(image_files)

    if async_mode:
        _start_async_task(job_id, 'barcode_scan', total, _do_barcode_scan,
                          double_sided, use_barcode, barcode_region)
        return jsonify({'started': True, 'task': 'barcode_scan', 'total': total})

    # 同步模式（兼容旧行为）
    result = _do_barcode_scan(job_id, lambda c, m: None, double_sided, use_barcode, barcode_region)
    return jsonify(result)


@app.route('/api/manual-match', methods=['POST'])
def manual_match():
    """人工匹配：为无法识别条码的图片指定学生"""
    job_id = request.json.get('job_id', '')
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    matches = request.json.get('matches', [])  # [{idx: N, student_id: '...'}]
    if not matches:
        return jsonify({'error': '没有匹配数据'}), 400

    job = jobs[job_id]
    if 'manual_matches' not in job:
        job['manual_matches'] = {}

    for m in matches:
        key = str(m.get('idx', m.get('pair_idx', '')))
        sid = m.get('student_id', '').strip()
        if key and sid:
            job['manual_matches'][key] = sid

    return jsonify({
        'ok': True,
        'count': len(job['manual_matches']),
    })


# ============================================================
#  位置保存 & 模板
# ============================================================

@app.route('/api/save-positions', methods=['POST'])
def save_positions():
    """保存分数打印位置坐标"""
    job_id = request.json.get('job_id', '')
    side = request.json.get('side', 'front')
    positions = request.json.get('positions', [])
    y_offset = request.json.get('y_offset', 0)
    x_offset = request.json.get('x_offset', 0)

    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    adjusted = []
    for p in positions:
        adjusted.append({
            'x': round(p['x'] + x_offset, 2),
            'y': round(p['y'] + y_offset, 2),
        })

    if side == 'front':
        jobs[job_id]['positions_front'] = adjusted
    else:
        jobs[job_id]['positions_back'] = adjusted

    return jsonify({
        'count': len(adjusted),
        'positions': adjusted,
    })


# ============================================================
#  处理
# ============================================================

@app.route('/api/process', methods=['POST'])
def process():
    """执行匹配和分数叠加处理。支持 async_mode 后台执行以显示进度条。"""
    job_id = request.json.get('job_id', '')
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    job = jobs[job_id]
    config = job.get('score_config')
    if not config:
        return jsonify({'error': '请先配置成绩文件'}), 400

    positions_front = job.get('positions_front', [])
    if not positions_front:
        return jsonify({'error': '请先标记分数打印位置'}), 400

    double_sided = request.json.get('double_sided', False)
    use_barcode = request.json.get('use_barcode', True)
    font_size = request.json.get('font_size', 80)
    create_class_folders = request.json.get('create_class_folders', True)
    async_mode = request.json.get('async', True)

    positions_back = job.get('positions_back', [])
    if double_sided and not positions_back:
        return jsonify({'error': '双面模式需要标记背面的分数位置'}), 400

    # 计算总工作量
    job_dir = get_job_dir(job_id)
    img_dir = job_dir / 'images'
    image_files = sorted([p for p in img_dir.iterdir() if _is_image_file(p)])
    total = len(image_files) // 2 if double_sided else len(image_files)

    if async_mode:
        _start_async_task(job_id, 'process', total, _do_process,
                          double_sided, use_barcode, font_size, create_class_folders)
        return jsonify({'started': True, 'task': 'process', 'total': total})

    # 同步模式（兼容旧行为）
    try:
        result = _do_process(job_id, lambda c, m: None, double_sided, use_barcode,
                             font_size, create_class_folders)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    job['matched'] = result['matched_count']
    job['unmatched_students'] = result['unmatched_students']
    job['unmatched_images'] = result['unmatched_images']
    job['output_files'] = result['output_files']
    job['status'] = 'done'

    output_dir = job_dir / 'output'
    return jsonify({
        'matched': result['matched_count'],
        'unmatched_students': len(result['unmatched_students']),
        'unmatched_images': len(result['unmatched_images']),
        'output_files': [str(Path(f).relative_to(output_dir)) for f in result['output_files'][:100]],
    })


# ============================================================
#  预览
# ============================================================

@app.route('/api/preview-result', methods=['POST'])
def preview_result():
    """生成预览图，用于检查位置和字号效果"""
    job_id = request.json.get('job_id', '')
    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    job = jobs[job_id]
    config = job.get('score_config')
    if not config:
        return jsonify({'error': '请先配置成绩文件'}), 400

    positions_front = job.get('positions_front', [])
    if not positions_front:
        return jsonify({'error': '请先在步骤标记分数位置'}), 400

    font_size = request.json.get('font_size', 80)
    double_sided = request.json.get('double_sided', False)
    positions_back = job.get('positions_back', [])

    rows = job['score_rows']
    parsed = _parse_scores_with_config(rows, config)
    if not parsed['students']:
        return jsonify({'error': '无学生数据'}), 400

    student = parsed['students'][0]
    full_scores = parsed['full_scores']
    student_scores = student['scores']

    img_dir = get_job_dir(job_id) / 'images'
    image_files = sorted([p for p in img_dir.iterdir() if _is_image_file(p)])
    if not image_files:
        return jsonify({'error': '请先上传答题卡图片'}), 400

    font = find_chinese_font(font_size)
    if font is None:
        return jsonify({'error': '无法加载字体'}), 500

    job_dir = get_job_dir(job_id)

    def _score_text(val, idx):
        fs = full_scores[idx] if idx < len(full_scores) else '?'
        return f'{val}/{fs}'

    # 正面预览
    n_front = len(positions_front)
    front_texts = [_score_text(student_scores[j], j) for j in range(min(n_front, len(student_scores)))]
    while len(front_texts) < n_front:
        front_texts.append('?/?')

    preview_img_name = request.json.get('image_name', '')
    if preview_img_name:
        preview_path = img_dir / preview_img_name
        if not preview_path.exists():
            preview_path = image_files[0]
    else:
        # 双面默认用第一张，单面用第一张
        preview_path = image_files[0]

    out_path = str(job_dir / 'preview_front.jpg')
    add_scores_to_image(
        str(preview_path),
        [p['x'] for p in positions_front],
        [p['y'] for p in positions_front],
        front_texts, font_size, out_path, font,
    )

    result = {
        'front_preview_url': '/api/preview-image/' + job_id + '/preview_front.jpg',
    }

    # 双面：生成背面预览
    if double_sided and positions_back:
        n_back = len(positions_back)
        back_texts = [_score_text(student_scores[n_front + j], n_front + j)
                      for j in range(min(n_back, len(student_scores) - n_front))]
        while len(back_texts) < n_back:
            back_texts.append('?/?')

        # 双面时用第二张图片（如有）作为背面预览
        back_img = image_files[1] if len(image_files) > 1 else image_files[0]
        back_out = str(job_dir / 'preview_back.jpg')
        add_scores_to_image(
            str(back_img),
            [p['x'] for p in positions_back],
            [p['y'] for p in positions_back],
            back_texts, font_size, back_out, font,
        )
        result['back_preview_url'] = '/api/preview-image/' + job_id + '/preview_back.jpg'

    return jsonify(result)


# ============================================================
#  保存 / 清理 / 预览图片 / 其它
# ============================================================

@app.route('/api/save-to-folder', methods=['POST'])
def save_to_folder():
    """将处理结果拷贝到用户指定文件夹，然后清理所有缓存"""
    job_id = request.json.get('job_id', '')
    target_path = (request.json.get('target_path') or '').strip()

    if not job_id or job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400
    if not target_path:
        return jsonify({'error': '请指定保存路径'}), 400

    job = jobs[job_id]
    output_dir = get_job_dir(job_id) / 'output'
    unmatched_dir = get_job_dir(job_id) / 'unmatched'

    if not output_dir.exists() or not any(output_dir.iterdir()):
        return jsonify({'error': '没有可保存的结果'}), 400

    target = Path(target_path)
    try:
        target.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return jsonify({'error': f'无法创建目标目录: {str(e)}'}), 400

    copied_files = 0
    errors = []

    match_dir = target / '已匹配'
    try:
        match_dir.mkdir(exist_ok=True)
        for root, dirs, files in os.walk(str(output_dir)):
            for fn in files:
                src = Path(root) / fn
                rel = Path(root).relative_to(str(output_dir))
                dst_dir = match_dir / rel
                dst_dir.mkdir(parents=True, exist_ok=True)
                try:
                    shutil.copy2(str(src), str(dst_dir / fn))
                    copied_files += 1
                except Exception as e:
                    errors.append(f'拷贝失败 {fn}: {str(e)}')
    except Exception as e:
        errors.append(f'创建"已匹配"目录失败: {str(e)}')

    if unmatched_dir.exists():
        unmatched_target = target / '无条码'
        try:
            for fn in os.listdir(str(unmatched_dir)):
                src = unmatched_dir / fn
                if src.is_file():
                    unmatched_target.mkdir(exist_ok=True)
                    try:
                        shutil.copy2(str(src), str(unmatched_target / fn))
                        copied_files += 1
                    except Exception as e:
                        errors.append(f'拷贝失败 {fn}: {str(e)}')
        except Exception as e:
            errors.append(f'拷贝"无条码"失败: {str(e)}')

    cleanup_data()
    jobs.pop(job_id, None)

    return jsonify({
        'copied_files': copied_files,
        'target_path': str(target),
        'errors': errors,
    })


@app.route('/api/output-file-list/<job_id>')
def output_file_list(job_id):
    """列出所有输出文件"""
    if job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    job_dir = get_job_dir(job_id)
    output_dir = job_dir / 'output'
    unmatched_dir = job_dir / 'unmatched'

    files = []

    if output_dir.exists():
        for root, dirs, fnames in os.walk(str(output_dir)):
            for fn in fnames:
                src = Path(root) / fn
                rel = str(src.relative_to(output_dir)).replace('\\', '/')
                files.append({
                    'path': '已匹配/' + rel,
                    'size': src.stat().st_size,
                    'type': 'matched',
                })

    if unmatched_dir.exists():
        for fn in os.listdir(str(unmatched_dir)):
            src = unmatched_dir / fn
            if src.is_file():
                files.append({
                    'path': '无条码/' + fn,
                    'size': src.stat().st_size,
                    'type': 'unmatched',
                })

    return jsonify({'files': files})


@app.route('/api/output-file/<job_id>/<path:filepath>')
def output_file_download(job_id, filepath):
    """下载单个输出文件"""
    if job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    job_dir = get_job_dir(job_id)

    parts = filepath.replace('\\', '/').split('/', 1)
    if len(parts) != 2:
        return jsonify({'error': '无效的文件路径'}), 400

    prefix, rest = parts
    if prefix == '已匹配':
        full_path = (job_dir / 'output' / rest).resolve()
    elif prefix == '无条码':
        full_path = (job_dir / 'unmatched' / rest).resolve()
    else:
        return jsonify({'error': '无效的文件路径'}), 400

    if not str(full_path).startswith(str(job_dir.resolve())):
        return jsonify({'error': '路径越界'}), 403

    if not full_path.exists() or not full_path.is_file():
        return jsonify({'error': '文件不存在'}), 404

    return send_file(str(full_path))


@app.route('/api/cleanup/<job_id>', methods=['POST'])
def cleanup_job(job_id):
    """清理指定 job 的缓存数据"""
    if job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    jobs.pop(job_id, None)
    cleanup_data()

    return jsonify({'ok': True})


@app.route('/api/preview-image/<job_id>/<filename>')
def preview_image(job_id, filename):
    """提供图片预览"""
    if job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    job_dir = get_job_dir(job_id)
    img_path = job_dir / 'images' / secure_filename(filename)
    if not img_path.exists():
        img_path = job_dir / secure_filename(filename)
    if not img_path.exists():
        return jsonify({'error': '图片不存在'}), 404

    return send_file(str(img_path))


@app.route('/api/image-list/<job_id>')
def image_list(job_id):
    """获取已上传的图片列表"""
    if job_id not in jobs:
        return jsonify({'error': '无效的任务ID'}), 400

    img_dir = get_job_dir(job_id) / 'images'
    images = sorted([p.name for p in img_dir.iterdir() if _is_image_file(p)])
    return jsonify({'images': images})


@app.route('/api/position-templates/save', methods=['POST'])
def save_position_template():
    """保存位置模板"""
    name = (request.json.get('name') or 'default').strip()
    positions_front = request.json.get('positions_front', [])
    positions_back = request.json.get('positions_back', [])
    y_offset = request.json.get('y_offset', 0)
    x_offset = request.json.get('x_offset', 0)

    if not positions_front:
        return jsonify({'error': '没有正面位置数据可保存'}), 400

    template_dir = TEMPLATES_POS_DIR
    template_dir.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r'[^\w\-]', '_', name)
    data = {
        'name': name,
        'positions_front': positions_front,
        'positions_back': positions_back,
        'y_offset': y_offset,
        'x_offset': x_offset,
    }
    with open(template_dir / f'{safe_name}.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return jsonify({'ok': True, 'name': name})


@app.route('/api/position-templates/list')
def list_position_templates():
    """列出已保存的位置模板"""
    template_dir = TEMPLATES_POS_DIR
    if not template_dir.exists():
        return jsonify({'templates': []})

    templates = []
    for p in sorted(template_dir.glob('*.json')):
        try:
            with open(p, 'r', encoding='utf-8') as f:
                data = json.load(f)
            templates.append({
                'filename': p.stem,
                'name': data.get('name', p.stem),
                'front_count': len(data.get('positions_front', [])),
                'back_count': len(data.get('positions_back', [])),
                'y_offset': data.get('y_offset', 0),
                'x_offset': data.get('x_offset', 0),
                'positions_front': data.get('positions_front', []),
                'positions_back': data.get('positions_back', []),
            })
        except Exception:
            continue
    return jsonify({'templates': templates})


@app.route('/api/position-templates/delete/<name>', methods=['POST'])
def delete_position_template(name):
    """删除位置模板"""
    template_dir = TEMPLATES_POS_DIR
    safe_name = re.sub(r'[^\w\-]', '_', name)
    path = template_dir / f'{safe_name}.json'
    if path.exists():
        path.unlink()
        return jsonify({'ok': True})
    return jsonify({'error': '模板不存在'}), 404


@app.route('/api/position-templates/clear-all', methods=['POST'])
def clear_all_position_templates():
    """清除所有位置模板"""
    template_dir = TEMPLATES_POS_DIR
    if template_dir.exists():
        count = 0
        for p in template_dir.glob('*.json'):
            p.unlink()
            count += 1
        return jsonify({'ok': True, 'deleted': count})
    return jsonify({'ok': True, 'deleted': 0})


@app.route('/api/dependencies')
def check_dependencies():
    """检查依赖库状态"""
    return jsonify({
        'pyzbar': HAS_PYZBAR,
        'pyzbar_functional': _check_pyzbar_functional(),
        'pymupdf': HAS_PYMUPDF,
        'openpyxl': HAS_OPENPYXL,
        'xlrd': HAS_XLRD,
        'pypinyin': HAS_PYPINYIN,
        'pillow': True,
    })


# ============================================================
#  启动入口
# ============================================================

if __name__ == '__main__':
    # PyInstaller 打包后子进程通过重新执行 exe 启动，必须用 freeze_support()
    # 让 multiprocessing 识别 worker 启动而不重复跑主程序逻辑
    multiprocessing.freeze_support()

    import sys
    port = 5000
    if len(sys.argv) > 1:
        try:
            port = int(sys.argv[1])
        except ValueError:
            pass

    print('=' * 50)
    print('  答题卡小题分匹配系统')
    print(f'  打开浏览器访问: http://localhost:{port}')
    print('=' * 50)

    app.run(host='0.0.0.0', port=port, debug=False)
